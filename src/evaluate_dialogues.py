import pandas as pd
from itertools import product

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
import json
import sys
import os
import platform
import sys
import time
from matplotlib import pyplot as plt
import json


class MissingEnvironmentVariable(Exception):
    pass







input_mode = "ASYNC GAZE+SPEECH" # Options: "ASYNC GAZE+SPEECH+SCENE", "SYNC GAZE+SPEECH+SCENE", 
                                       # "ASYNC GAZE+SPEECH", "SYNC GAZE+SPEECH", "SPEECH+SCENE", "GAZE+SCENE"
scenario = "Drink"
data_dir = "/hri/localdisk/emende/generated_dialogues"+"/"+scenario+"/6_users"

data_file = input_mode+"_"+scenario+".xlsx"
data_file_path = os.path.join(data_dir,data_file)




config_files_dict = {
    "ASYNC GAZE+SPEECH+SCENE": "gpt_gaze_speech_scene_config",
    "SYNC GAZE+SPEECH+SCENE": "gpt_sync_gaze_speech_scene_config",
    "ASYNC GAZE+SPEECH": "gpt_gaze_speech_config",
    "SYNC GAZE+SPEECH": "gpt_sync_gaze_speech_config",
    "SPEECH+SCENE": "gpt_speech_scene_config",
    "GAZE+SCENE": "gpt_gaze_scene_config"}

columns = {
    "INTERACTION": "A",
    "USER": "B",
    "RUN": "C",
    "DESIRED SPEECH": "D",
    "DESIRED GAZE": "E",
    "SPEECH": "F",
    "GAZE": "G",
    "SYNCHRONIZED": "H",
    "QUERY_OBJECTS": "I",
    "QUERY_AGENTS": "J",
    "DESIRED REASONING": "K",
    "DESIRED SPEAKING": "L",
    "REASONING": "M",
    "SPEAKING": "N",
    "REQUIRED OBJECTS": "O", 
}

columns_width = {
    "INTERACTION": 15,
    "USER": 10,
    "RUN": 10,
    "DESIRED SPEECH": 30,
    "DESIRED GAZE": 30,
    "SPEECH": 30,
    "GAZE": 50,
    "SYNCHRONIZED": 50,
    "QUERY_OBJECTS": 30,
    "QUERY_AGENTS": 30,
    "DESIRED REASONING": 50,
    "DESIRED SPEAKING": 50,
    "REASONING": 50,
    "SPEAKING": 50,
    "REQUIRED OBJECTS": 50,
}

excluded_objects = ['hand_left_robot', 'hand_right_robot']


correct_interactions_breakfast = [0,0,0]
number_interactions_breakfast = [0,0,0]

correct_interactions_drink = [0,0,0]
number_interactions_drink = [0,0,0]

#Read row interaction 

if os.path.exists(data_file_path):
    wb = load_workbook(data_file_path)
    ws = wb["Generated Dialogues Scheme"]    

    for row in range(6, ws.max_row + 1):  # Skip the header row
        required_objects = ws[f"{columns['REQUIRED OBJECTS']}{row}"].value
        if required_objects:
            # Split the string into a list of objects
            required_objects_list = [obj.strip() for obj in required_objects.split(",")]
        else:
            required_objects_list = []

        if scenario == "Breakfast":
            if ws[f"{columns['INTERACTION']}{row}"].value == 1:
                number_interactions_breakfast[0] += 1         
                if "bowl" in required_objects_list and "box_of_cereal" in required_objects_list:
                    correct_interactions_breakfast[0] += 1
                        
            if ws[f"{columns['INTERACTION']}{row}"].value == 2:
                number_interactions_breakfast[1] += 1         
                if "bottle_of_milk" in required_objects_list and "bottle_of_orange_juice" not in required_objects_list:
                    correct_interactions_breakfast[1] += 1
                
            if ws[f"{columns['INTERACTION']}{row}"].value == 3:
                number_interactions_breakfast[2] += 1         
                if "small_bowl" in required_objects_list:
                    correct_interactions_breakfast[2] += 1
        if scenario == "Drink":
            if ws[f"{columns['INTERACTION']}{row}"].value == 1:
                number_interactions_drink[0] += 1 
                
                if any(obj == "bottle_of_cola" for obj in required_objects_list) and not any(obj == "bottle_of_cola_zero" for obj in required_objects_list):        
                    print(required_objects_list)
                    correct_interactions_drink[0] += 1
                    
            elif ws[f"{columns['INTERACTION']}{row}"].value == 2:
                number_interactions_drink[1] += 1         
                if "glass_red" in required_objects_list and "glass_blue" not in required_objects_list:
                    correct_interactions_drink[1] += 1
                
            elif ws[f"{columns['INTERACTION']}{row}"].value == 3:
                number_interactions_drink[2] += 1         
                if "bowl" in required_objects_list:
                    correct_interactions_drink[2] += 1

if scenario == 'Breakfast':
    print(number_interactions_breakfast) 
    print(correct_interactions_breakfast)
if scenario == 'Drink':
    print(number_interactions_drink)
    print(correct_interactions_drink)



