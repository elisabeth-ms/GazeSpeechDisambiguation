import pandas as pd
from itertools import product

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
import json
import pyGaze
import sys
import os
import platform
import sys
import time
import py_LLM_handler
from matplotlib import pyplot as plt
sys.path.append(os.path.abspath("/hri/localdisk/emende/AttentiveSupport/src"))
from function_analyzer import FunctionAnalyzer

import json
from openai.types.chat.chat_completion import ChatCompletionMessage

if platform.system() == "Linux":
    sys.path.append("lib")
elif platform.system() == "Windows":
    sys.path.append("bin")

from pyAffaction import *
sys.path.append(os.path.abspath("/hri/localdisk/emende/AttentiveSupport/src/Smile/src/AffAction/python"))
import pyGaze

class MissingEnvironmentVariable(Exception):
    pass


if "OPENAI_API_KEY" not in os.environ:
    raise MissingEnvironmentVariable(
        "Please set an environment variable with your OPENAI_API_KEY. "
        "See https://help.openai.com/en/articles/5112595-best-practices-for-api-key-safety"
    )



SIM = None
print_emojis = True  
speech_directory_path = None
gpt_responses_path = None
filtered_gaze_data_directory_path = None
recordTransformationsEnabled = None


main_dir = "/hri/localdisk/emende/interaction_recordings/users"
input_mode = "ASYNC GAZE+SPEECH" # Options: "ASYNC GAZE+SPEECH+SCENE", "SYNC GAZE+SPEECH+SCENE", 
                                       # "ASYNC GAZE+SPEECH", "SYNC GAZE+SPEECH", "SPEECH+SCENE", "GAZE+SCENE"
scenario = "Breakfast"
task = 3

save_dir = "/hri/localdisk/emende/isolated_interactions"+"/"+scenario

output_file = input_mode+"_"+scenario+"_task"+str(task)+".xlsx"
output_file_path = os.path.join(save_dir,output_file)




config_files_dict = {
    "ASYNC GAZE+SPEECH+SCENE": "gpt_gaze_speech_scene_config",
    "SYNC GAZE+SPEECH+SCENE": "gpt_sync_gaze_speech_scene_config",
    "ASYNC GAZE+SPEECH": "gpt_gaze_speech_config",
    "SYNC GAZE+SPEECH": "gpt_sync_gaze_speech_config",
    "SPEECH+SCENE": "gpt_speech_scene_config",
    "GAZE+SCENE": "gpt_gaze_scene_config"}

columns = {
    "INTERACTION": "A",
    "USER": "B",
    "RUN": "C",
    "DESIRED SPEECH": "D",
    "DESIRED GAZE": "E",
    "SPEECH": "F",
    "GAZE": "G",
    "SYNCHRONIZED": "H",
    "QUERY_OBJECTS": "I",
    "QUERY_AGENTS": "J",
    "DESIRED REASONING": "K",
    "DESIRED SPEAKING": "L",
    "REASONING": "M",
    "SPEAKING": "N",
    "REQUIRED OBJECTS": "O", 
}

columns_width = {
    "INTERACTION": 15,
    "USER": 10,
    "RUN": 10,
    "DESIRED SPEECH": 30,
    "DESIRED GAZE": 30,
    "SPEECH": 30,
    "GAZE": 50,
    "SYNCHRONIZED": 50,
    "QUERY_OBJECTS": 30,
    "QUERY_AGENTS": 30,
    "DESIRED REASONING": 50,
    "DESIRED SPEAKING": 50,
    "REASONING": 50,
    "SPEAKING": 50,
    "REQUIRED OBJECTS": 50,
}

excluded_objects = ['hand_left_robot', 'hand_right_robot']



# Function to load speech data from a JSON file
def load_speech_data(speech_data_file):
    try:
        # Check if the file exists
        if not os.path.exists(speech_data_file):
            print(f"File not found: {speech_data_file}")
            return None

        # Open and load the JSON file
        with open(speech_data_file, 'r') as f:
            speech_data = json.load(f)
        return speech_data

    except Exception as e:
        print(f"Error loading file {speech_data_file}: {e}")
        return None

def load_info_from_json(json_file):
    with open(json_file, 'r') as f:
        data = json.load(f)
        print(data)
    return data


# Function to load gaze data from a JSON file
def load_gaze_data(gaze_data_file):
    try:
        if not os.path.exists(gaze_data_file):
            print(f"File not found: {gaze_data_file}")
            return None
        with open(gaze_data_file, 'r') as f:
            gaze_data = json.load(f)
        return gaze_data
    except Exception as e:
        print(f"Error loading file {gaze_data_file}: {e}")
        return None


if not os.path.exists(output_file_path):
    

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Isolated Interactions"

    # Add hierarchical headers
    ws.merge_cells("A1:J1")  # Merge for "GENERATED DIALOGUES"
    ws["A1"] = "ISOLATED INTERACTIONS"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:D2")  # Merge for "INPUT MODE"
    ws["A2"] = "INPUT MODE"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("E2:J2")  
    ws["E2"] = input_mode
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:D3")  # Merge for "SCENARIO"
    ws["A3"] = "SCENARIO"
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")


    ws.merge_cells("E3:F3")  
    ws["E3"] = scenario
    ws["E3"].alignment = Alignment(horizontal="center", vertical="center")

    ws["G3"] = "TASK"
    ws["G3"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["H3"] = task
    ws["H3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sub-headers for other columns
    start_row = 4
    end_row = 5
    ws.merge_cells(f"{columns['INTERACTION']}{start_row}:{columns['INTERACTION']}{end_row}")
    ws[f"{columns['INTERACTION']}{start_row}"] = 'INTERACTION'

    ws.merge_cells(f"{columns['USER']}{start_row}:{columns['USER']}{end_row}")
    ws[f"{columns['USER']}{start_row}"] = 'USER'

    ws.merge_cells(f"{columns['RUN']}{start_row}:{columns['RUN']}{end_row}")
    ws[f"{columns['RUN']}{start_row}"] = 'RUN'

    ws.merge_cells(f"{columns['DESIRED SPEECH']}{start_row}:{columns['DESIRED SPEECH']}{end_row}")
    ws[f"{columns['DESIRED SPEECH']}{start_row}"] = 'DESIRED SPEECH'
    ws[f"{columns['DESIRED SPEECH']}{start_row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"{columns['DESIRED GAZE']}{start_row}:{columns['DESIRED GAZE']}{end_row}")
    ws[f"{columns['DESIRED GAZE']}{start_row}"] = 'DESIRED GAZE'
    ws[f"{columns['DESIRED GAZE']}{start_row}"].alignment = Alignment(horizontal="center", vertical="center")



    ws.merge_cells(f"{columns['SPEECH']}{start_row}:{columns['SYNCHRONIZED']}{start_row}")
    ws[f"{columns['SPEECH']}{start_row}"] = 'INPUT'
    ws[f"{columns['SPEECH']}{start_row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"{columns['SPEECH']}{end_row}"] = 'SPEECH'
    ws[f"{columns['GAZE']}{end_row}"] = 'GAZE'
    ws[f"{columns['SYNCHRONIZED']}{end_row}"] = 'SYNCHRONIZED'

    ws.merge_cells(f"{columns['QUERY_OBJECTS']}{start_row}:{columns['QUERY_OBJECTS']}{end_row}")
    ws[f"{columns['QUERY_OBJECTS']}{start_row}"] = 'QUERY_OBJECTS'


    ws.merge_cells(f"{columns['QUERY_AGENTS']}{start_row}:{columns['QUERY_AGENTS']}{end_row}")
    ws[f"{columns['QUERY_AGENTS']}{start_row}"] = 'QUERY_AGENTS'

    ws.merge_cells(f"{columns['DESIRED REASONING']}{start_row}:{columns['DESIRED REASONING']}{end_row}")
    ws[f"{columns['DESIRED REASONING']}{start_row}"] = 'DESIRED REASONING'

    ws.merge_cells(f"{columns['DESIRED SPEAKING']}{start_row}:{columns['DESIRED SPEAKING']}{end_row}")
    ws[f"{columns['DESIRED SPEAKING']}{start_row}"] = 'DESIRED SPEAKING'

    ws.merge_cells(f"{columns['REASONING']}{start_row}:{columns['REASONING']}{end_row}")
    ws[f"{columns['REASONING']}{start_row}"] = 'REASONING'

    ws.merge_cells(f"{columns['SPEAKING']}{start_row}:{columns['SPEAKING']}{end_row}")
    ws[f"{columns['SPEAKING']}{start_row}"] = 'SPEAKING'
    
    ws.merge_cells(f"{columns['REQUIRED OBJECTS']}{start_row}:{columns['REQUIRED OBJECTS']}{end_row}")
    ws[f"{columns['REQUIRED OBJECTS']}{start_row}"] = 'REQUIRED OBJECTS'

    for key, value in columns.items():
        ws.column_dimensions[value].width = columns_width[key]
else:
    wb = load_workbook(output_file_path)
    ws = wb["Isolated Interactions"]    

# Define the users and interactions
users = [1,2,3,4,5,6]


row = 6

print_emojis = True
user_name = "Elisabeth"
llm_handler = py_LLM_handler.LLMHandler(config_module=config_files_dict[input_mode])
SIM = llm_handler.get_simulation()
SIM.run()

runs_per_user = 10

for user in users:
    
    for run in range(runs_per_user):
        print(f"User {user}, Run {run + 1}")
        ws[f"{columns['INTERACTION']}{row}"] = task
        ws[f"{columns['USER']}{row}"] = user
        cell_value = ws[f"{columns['RUN']}{row}"].value

        # Debugging: print the raw value and its type
        # print(f"Row {row}, Column RUN, Raw Value: {cell_value}, Type: {type(cell_value)}")
        if cell_value != None:
            print(f"Row {row}, Column RUN, Value: {cell_value}")
            row += 1
            continue
        

        user_dir = f"{main_dir}/user{user}"
        print(user_dir)
            
        user_scenario_dir = f"{user_dir}/{scenario}"
            
        print(user_scenario_dir)
            
        user_interaction_dir = f"{user_scenario_dir}/{scenario}{task}"
            
        print(user_interaction_dir)
            
            
            
        info_data = load_info_from_json(f"{user_interaction_dir}/info.json")
            
            
            
        speech_file = f"{user_interaction_dir}/speech_data/Elisabeth.json"
            
        speech_data = load_speech_data(speech_file)
            
        ws[f"{columns['DESIRED SPEECH']}{row}"] = info_data["speech"]
            
        ws[f"{columns['DESIRED GAZE']}{row}"] = info_data["gaze"]
            
        ws[f"{columns['DESIRED REASONING']}{row}"] = info_data["desired_reason"]
            
        ws[f"{columns['DESIRED SPEAKING']}{row}"] = info_data["desired_speaking"]
            
            
        speech_input = ""
        if speech_data:
            speech_input = " ".join(speech_data["transcript"])            
            start_time = speech_data["listening_start_time"]
                
            word_data = [(word_info['word'], word_info['start_time'], word_info['end_time']) for word_info in speech_data['words']]

            
        gaze_data_file = f"{user_interaction_dir}/raw_gaze_data/Elisabeth.json"
        # Load pre-recorded gaze and speech data
        user_raw_gaze_data = load_gaze_data(gaze_data_file)
            
            
        gaze_history = ""
        if user_raw_gaze_data:
            gaze_history, objects_timestamps = pyGaze.compute_list_closest_objects_gaze_history(user_raw_gaze_data["gaze_data"], start_time, 15.0,8.0, 8.0, excluded_objects, 5.0, 0.5, 0.08)
            
        ws[f"{columns['RUN']}{row}"] = run + 1
        ws[f"{columns['SPEECH']}{row}"] = speech_input
        ws[f"{columns['GAZE']}{row}"] = str(gaze_history)
            
            
        if input_mode == "SPEECH+SCENE" and speech_input:
            ws[f"{columns['GAZE']}{row}"] = "Not Applicable"
            print(f"{llm_handler._user_speech_emojis if print_emojis else ''}{speech_input}")
            llm_handler.play_with_functions_synchronized(speech_input, person_name=user_name)
            for message in llm_handler.messages_current_call:

                if not isinstance(message, ChatCompletionMessage):
                    print(message)
                    print("-----------------------------------")
                    if message.get("name") == "query_objects":
                        ws[f"{columns['QUERY_OBJECTS']}{row}"] = message.get("content")
                    if message.get("name") == "query_agents":
                        ws[f"{columns['QUERY_AGENTS']}{row}"] = message.get("content")
                    if message.get("name") == "reasoning":
                        reasoning_msg = message.get("content")
                        reasoning_msg = reasoning_msg.replace("You are about to take the following action: ", "", 1).strip() 
                        ws[f"{columns['REASONING']}{row}"] = reasoning_msg
                    if message.get("name") == "speak":
                        speak_msg = message.get("content")
                        speak_msg = speak_msg.replace("You said to Elisabeth: ", "", 1).strip()
                        ws[f"{columns['SPEAKING']}{row}"] = speak_msg
                    if message.get("name") == "required_objects":
                        required_objects_msg = message.get("content")
                        required_objects_msg = required_objects_msg.replace("The objects required by the user are: ", "", 1).strip()
                        ws[f"{columns['REQUIRED OBJECTS']}{row}"] = required_objects_msg

            llm_handler.messages_current_call = []

        if input_mode == "GAZE+SCENE" and gaze_history:
            ws[f"{columns['SPEECH']}{row}"] = "Not Applicable"
            print(f"{llm_handler._user_gaze_emojis if print_emojis else ''}{gaze_history}")
            llm_handler.play_with_functions_synchronized(gaze_history, person_name=user_name)
                
            for message in llm_handler.messages_current_call:

                if not isinstance(message, ChatCompletionMessage):
                    print(message)
                    print("-----------------------------------")
                    if message.get("name") == "query_objects":
                        ws[f"{columns['QUERY_OBJECTS']}{row}"] = message.get("content")
                    if message.get("name") == "query_agents":
                        ws[f"{columns['QUERY_AGENTS']}{row}"] = message.get("content")
                    if message.get("name") == "reasoning":
                        reasoning_msg = message.get("content")
                        reasoning_msg = reasoning_msg.replace("You are about to take the following action: ", "", 1).strip() 
                        ws[f"{columns['REASONING']}{row}"] = reasoning_msg
                    if message.get("name") == "speak":
                        speak_msg = message.get("content")
                        speak_msg = speak_msg.replace("You said to Elisabeth: ", "", 1).strip()
                        ws[f"{columns['SPEAKING']}{row}"] = speak_msg
                    if message.get("name") == "required_objects":
                        required_objects_msg = message.get("content")
                        required_objects_msg = required_objects_msg.replace("The objects required by the user are: ", "", 1).strip()
                        ws[f"{columns['REQUIRED OBJECTS']}{row}"] = required_objects_msg

            llm_handler.messages_current_call = []
            
        if not input_mode.startswith("SYNC"):
            ws[f"{columns['SYNCHRONIZED']}{row}"] = "Not Applicable"    
            
        if speech_input and gaze_history:
      
            if input_mode == "ASYNC GAZE+SPEECH+SCENE" or input_mode == "ASYNC GAZE+SPEECH":
                        
                print(f"{llm_handler._user_speech_emojis if print_emojis else ''}{speech_input}")
                print(f"{llm_handler._user_gaze_emojis if print_emojis else ''}{gaze_history}")
                llm_handler.play_with_functions_gaze_history_speech(speech_input=speech_input, gaze_history=gaze_history, person_name=user_name)
                print("LAST CALL RESPONSES: ")
                for message in llm_handler.messages_current_call:

                    if not isinstance(message, ChatCompletionMessage):
                        print(message)
                        print("-----------------------------------")
                        if message.get("name") == "query_objects":
                            ws[f"{columns['QUERY_OBJECTS']}{row}"] = message.get("content")
                        if message.get("name") == "query_agents":
                            ws[f"{columns['QUERY_AGENTS']}{row}"] = message.get("content")
                        if message.get("name") == "reasoning":
                            reasoning_msg = message.get("content")
                            reasoning_msg = reasoning_msg.replace("You are about to take the following action: ", "", 1).strip() 
                            ws[f"{columns['REASONING']}{row}"] = reasoning_msg
                        if message.get("name") == "speak":
                            speak_msg = message.get("content")
                            speak_msg = speak_msg.replace("You said to Elisabeth: ", "", 1).strip()
                            ws[f"{columns['SPEAKING']}{row}"] = speak_msg
                        if message.get("name") == "required_objects":
                            required_objects_msg = message.get("content")
                            required_objects_msg = required_objects_msg.replace("The objects required by the user are: ", "", 1).strip()
                            ws[f"{columns['REQUIRED OBJECTS']}{row}"] = required_objects_msg

                llm_handler.messages_current_call = []
            elif input_mode == "SYNC GAZE+SPEECH+SCENE" or input_mode == "SYNC GAZE+SPEECH":
                input_data = pyGaze.merge_gaze_word_intervals(objects_timestamps, word_data)
                ws[f"{columns['SYNCHRONIZED']}{row}"] = str(input_data)
                ws[f"{columns['GAZE']}{row}"] = str(objects_timestamps)
                print(f"{llm_handler._user_emojis if print_emojis else ''}{input_data}")
                print(f"{llm_handler._user_speech_emojis if print_emojis else ''}{speech_input}")
                print(f"{llm_handler._user_gaze_emojis if print_emojis else ''}{objects_timestamps}")
                llm_handler.play_with_functions_synchronized(input=input_data, person_name=user_name)
                print("LAST CALL RESPONSES: ")
                for message in llm_handler.messages_current_call:
                        
                    if not isinstance(message, ChatCompletionMessage):
                        print(message)
                        print("-----------------------------------")
                        if message.get("name") == "query_objects":
                            ws[f"{columns['QUERY_OBJECTS']}{row}"] = message.get("content")
                        if message.get("name") == "query_agents":
                            ws[f"{columns['QUERY_AGENTS']}{row}"] = message.get("content")
                        if message.get("name") == "reasoning":
                            reasoning_msg = message.get("content")
                            reasoning_msg = reasoning_msg.replace("You are about to take the following action: ", "", 1).strip() 
                            ws[f"{columns['REASONING']}{row}"] = reasoning_msg
                        if message.get("name") == "speak":
                            speak_msg = message.get("content")
                            speak_msg = speak_msg.replace("You said to Elisabeth: ", "", 1).strip()
                            ws[f"{columns['SPEAKING']}{row}"] = speak_msg
                        if message.get("name") == "required_objects":
                            required_objects_msg = message.get("content")
                            required_objects_msg = required_objects_msg.replace("The objects required by the user are: ", "", 1).strip()
                            ws[f"{columns['REQUIRED OBJECTS']}{row}"] = required_objects_msg

                llm_handler.messages_current_call = []


            row += 1
        # if row >=10:
        #     break
        print("New dialogue show we reset the LLM handler")
        llm_handler.reset()


    



wb.save(output_file_path)

print(f"Excel file '{output_file}' created successfully!")
